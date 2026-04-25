"""
Phase 2c — Export to Excel

Input:  data/unicorn_companies_clean.csv
Output: excel/unicorn_analysis.xlsx  (multi-sheet workbook)

Run:  python3 notebook/excel_export.py | tee report/excelReport.txt
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_GREEN  = "1A3C34"
MID_GREEN   = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
ACCENT      = "52B788"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F8F9FA"
MED_GRAY    = "DEE2E6"

# ── Helper: style a header row ─────────────────────────────────────────────────
def style_header(ws, row=1, color=DARK_GREEN):
    for cell in ws[row]:
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.font      = Font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="medium", color=ACCENT),
        )

# ── Helper: auto-fit column widths ────────────────────────────────────────────
def autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(max_len + 4, min_w), max_w)

# ── Helper: zebra-stripe data rows ───────────────────────────────────────────
def zebra(ws, start_row=2, col_count=None):
    col_count = col_count or ws.max_column
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        fill_color = LIGHT_GREEN if i % 2 == 0 else WHITE
        for cell in row[:col_count]:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(vertical="center")

# ── 1. Load data ──────────────────────────────────────────────────────────────
df = pd.read_csv("data/unicorn_companies_clean.csv", parse_dates=["date_joined"])
print(f"Loaded {len(df):,} rows")

# ── 2. Build summary DataFrames ───────────────────────────────────────────────
industry_summary = (
    df.groupby("industry")
    .agg(
        companies         = ("company",          "count"),
        avg_valuation_b   = ("valuation_b",      "mean"),
        median_valuation_b= ("valuation_b",      "median"),
        total_funding_b   = ("funding_b",        "sum"),
        avg_years_to_uni  = ("years_to_unicorn", "mean"),
        high_value_pct    = ("is_high_value",    "mean"),
    )
    .reset_index()
    .sort_values("companies", ascending=False)
    .round(2)
)
industry_summary["high_value_pct"] = (industry_summary["high_value_pct"] * 100).round(1)
industry_summary.columns = [
    "Industry", "# Companies", "Avg Valuation ($B)",
    "Median Valuation ($B)", "Total Funding ($B)",
    "Avg Years to Unicorn", "% High Value (≥$10B)",
]

country_summary = (
    df.groupby(["country", "continent"])
    .agg(
        companies       = ("company",     "count"),
        avg_valuation_b = ("valuation_b", "mean"),
        total_funding_b = ("funding_b",   "sum"),
    )
    .reset_index()
    .sort_values("companies", ascending=False)
    .head(30)
    .round(2)
)
country_summary.columns = [
    "Country", "Continent", "# Companies",
    "Avg Valuation ($B)", "Total Funding ($B)",
]

continent_summary = (
    df.groupby("continent")
    .agg(
        companies         = ("company",          "count"),
        avg_valuation_b   = ("valuation_b",      "mean"),
        total_funding_b   = ("funding_b",        "sum"),
        avg_years_to_uni  = ("years_to_unicorn", "mean"),
        high_value_count  = ("is_high_value",    "sum"),
    )
    .reset_index()
    .sort_values("companies", ascending=False)
    .round(2)
)
continent_summary.columns = [
    "Continent", "# Companies", "Avg Valuation ($B)",
    "Total Funding ($B)", "Avg Years to Unicorn", "# High Value (≥$10B)",
]

yearly_trends = (
    df.groupby("year_joined")
    .agg(
        new_unicorns    = ("company",     "count"),
        avg_valuation_b = ("valuation_b", "mean"),
        total_funding_b = ("funding_b",   "sum"),
    )
    .reset_index()
    .dropna(subset=["year_joined"])
    .sort_values("year_joined")
    .round(2)
)
yearly_trends["year_joined"] = yearly_trends["year_joined"].astype(int)
yearly_trends.columns = [
    "Year", "New Unicorns", "Avg Valuation ($B)", "Total Funding ($B)",
]

# ── 3. Write to Excel ─────────────────────────────────────────────────────────
output_path = "excel/unicorn_analysis.xlsx"

# Select display columns for the raw data sheet
raw_cols = [
    "company", "valuation_b", "funding_b", "industry",
    "country", "continent", "year_founded", "year_joined",
    "years_to_unicorn", "funding_efficiency", "is_high_value",
]
raw_display = df[raw_cols].copy()
raw_display.columns = [
    "Company", "Valuation ($B)", "Funding ($B)", "Industry",
    "Country", "Continent", "Year Founded", "Year Joined",
    "Years to Unicorn", "Funding Efficiency", "High Value?",
]

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    raw_display.to_excel(writer,       sheet_name="Raw Data",          index=False)
    industry_summary.to_excel(writer,  sheet_name="Industry Summary",  index=False)
    country_summary.to_excel(writer,   sheet_name="Country Summary",   index=False)
    continent_summary.to_excel(writer, sheet_name="Continent Summary", index=False)
    yearly_trends.to_excel(writer,     sheet_name="Yearly Trends",     index=False)

print("Base sheets written.")

# ── 4. Style workbook ─────────────────────────────────────────────────────────
wb = load_workbook(output_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"          # freeze header row
    ws.sheet_view.showGridLines = False

    # Header colour: darker for raw data, accent for summaries
    hdr_color = DARK_GREEN if sheet_name == "Raw Data" else MID_GREEN
    style_header(ws, row=1, color=hdr_color)
    zebra(ws, start_row=2)
    autofit(ws)

    # Row height
    ws.row_dimensions[1].height = 28
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

print("Styling applied.")

# ── 5. Add charts ─────────────────────────────────────────────────────────────

# Chart 1: Bar chart — unicorns per industry (Industry Summary sheet)
ws_ind = wb["Industry Summary"]
n_ind = len(industry_summary) + 1

bar = BarChart()
bar.type        = "bar"          # horizontal
bar.grouping    = "clustered"
bar.title       = "Unicorn companies by industry"
bar.y_axis.title = "Industry"
bar.x_axis.title = "# Companies"
bar.style       = 10
bar.width       = 22
bar.height      = 14

cats = Reference(ws_ind, min_col=1, min_row=2, max_row=n_ind)
data = Reference(ws_ind, min_col=2, min_row=1, max_row=n_ind)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = ACCENT

ws_ind.add_chart(bar, "I2")

# Chart 2: Line chart — new unicorns per year (Yearly Trends sheet)
ws_yr = wb["Yearly Trends"]
n_yr  = len(yearly_trends) + 1

line = LineChart()
line.title        = "New unicorns per year"
line.y_axis.title = "Count"
line.x_axis.title = "Year"
line.style        = 10
line.width        = 22
line.height       = 14

yr_cats = Reference(ws_yr, min_col=1, min_row=2, max_row=n_yr)
yr_data = Reference(ws_yr, min_col=2, min_row=1, max_row=n_yr)
line.add_data(yr_data, titles_from_data=True)
line.set_categories(yr_cats)
line.series[0].graphicalProperties.line.solidFill    = MID_GREEN
line.series[0].graphicalProperties.line.width        = 25000   # EMUs ~2pt

ws_yr.add_chart(line, "F2")

# ── 6. Save ───────────────────────────────────────────────────────────────────
wb.save(output_path)
print(f"\nSaved → {output_path}")
print("Sheets: Raw Data | Industry Summary | Country Summary | "
      "Continent Summary | Yearly Trends")
print("Charts: bar chart (industry) + line chart (yearly trends)")
"""
Phase 3, Step 3 — Excel Pivot Tables
Unicorn Companies Project
------------------------------------------------
Input:  data/unicorn_companies_clean.csv
Output: excel/unicorn_pivot_tables.xlsx  (5 pivot sheets + formatting)

Run:  python3 notebook/pivot_tables.py

Pivot tables:
  1. Industry × Continent  — company count matrix
  2. Valuation tiers       — how many companies fall in each bracket
  3. Decade × Industry     — founding decade breakdown
  4. Country league table  — rank, count, avg val, funding efficiency
  5. Yearly growth         — YoY new unicorns and cumulative total
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_GREEN   = "1A3C34"
MID_GREEN    = "2D6A4F"
ACCENT_GREEN = "52B788"
LIGHT_GREEN  = "D8F3DC"
VERY_LIGHT   = "F0FAF4"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F8F9FA"
MED_GRAY     = "DEE2E6"
TEXT_DARK    = "212529"

OUTPUT_PATH = "excel/unicorn_pivot_tables.xlsx"

# ── Load data ─────────────────────────────────────────────────────────────────
df = pd.read_csv("data/unicorn_companies_clean.csv", parse_dates=["date_joined"])
print(f"Loaded {len(df):,} rows\n")

# Normalise the duplicate AI label
df["industry"] = df["industry"].replace(
    "Artificial Intelligence", "Artificial intelligence"
)

# ── Helper: cell styling ──────────────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color=TEXT_DARK, size=10):
    return Font(bold=bold, color=color, size=size)

def make_border(style="thin", color=MED_GRAY):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, n_cols, row=3, row_num=None, bg=DARK_GREEN, font_size=11):
    row_num = row_num if row_num is not None else row
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = make_fill(bg)
        cell.font      = make_font(bold=True, color=WHITE, size=font_size)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = make_border("medium", ACCENT_GREEN)
    ws.row_dimensions[row_num].height = 28


def style_index_col(ws, start_row, end_row, col=1, bg=MID_GREEN):
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = make_fill(bg)
        cell.font      = make_font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(vertical="center")

def zebra_rows(ws, start_row, end_row, start_col, end_col):
    for i, row in enumerate(range(start_row, end_row + 1)):
        bg = LIGHT_GREEN if i % 2 == 0 else WHITE
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE):
                cell.fill = make_fill(bg)
            cell.alignment = Alignment(vertical="center", horizontal="center")

def autofit(ws, min_w=10, max_w=35):
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(max_len + 3, min_w), max_w)

def add_title(ws, text, row=1):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font      = make_font(bold=True, color=DARK_GREEN, size=13)
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24

# ── Build DataFrames ──────────────────────────────────────────────────────────

# 1. Industry × Continent matrix
pivot1 = pd.pivot_table(
    df, values="company", index="industry",
    columns="continent", aggfunc="count", fill_value=0
)
pivot1["TOTAL"] = pivot1.sum(axis=1)
pivot1 = pivot1.sort_values("TOTAL", ascending=False)
pivot1.loc["TOTAL"] = pivot1.sum()

# 2. Valuation tiers
bins   = [0, 1, 2, 5, 10, 20, 50, float("inf")]
labels = ["$1B", "$1–2B", "$2–5B", "$5–10B", "$10–20B", "$20–50B", "$50B+"]
df["val_tier"] = pd.cut(df["valuation_b"], bins=bins, labels=labels, right=True)
pivot2 = (df.groupby("val_tier", observed=True)
          .agg(
              companies    = ("company",          "count"),
              avg_val      = ("valuation_b",      "mean"),
              total_val    = ("valuation_b",      "sum"),
              avg_funding  = ("funding_b",        "mean"),
              avg_yrs      = ("years_to_unicorn", "mean"),
          )
          .round(2)
          .reset_index())
pivot2.columns = [
    "Valuation tier", "# Companies", "Avg valuation ($B)",
    "Total valuation ($B)", "Avg funding ($B)", "Avg years to unicorn",
]

# 3. Decade × Industry
df["decade"] = (df["year_founded"] // 10 * 10).astype("Int64").astype(str) + "s"
pivot3 = pd.pivot_table(
    df, values="company", index="decade",
    columns="industry", aggfunc="count", fill_value=0
)
pivot3["TOTAL"] = pivot3.sum(axis=1)
pivot3 = pivot3.sort_index()

# 4. Country league table (top 25)
pivot4 = (df.groupby(["country", "continent"])
          .agg(
              unicorns         = ("company",          "count"),
              avg_valuation    = ("valuation_b",      "mean"),
              total_valuation  = ("valuation_b",      "sum"),
              avg_funding      = ("funding_b",        "mean"),
              avg_yrs          = ("years_to_unicorn", "mean"),
              high_value       = ("is_high_value",    "sum"),
          )
          .reset_index()
          .sort_values("unicorns", ascending=False)
          .head(25)
          .round(2))
pivot4.insert(0, "Rank", range(1, len(pivot4) + 1))
pivot4.columns = [
    "Rank", "Country", "Continent", "# Unicorns", "Avg Valuation ($B)",
    "Total Valuation ($B)", "Avg Funding ($B)",
    "Avg Years to Unicorn", "# High Value (≥$10B)",
]

# 5. Yearly growth
pivot5 = (df.groupby("year_joined")
          .agg(new_unicorns=("company", "count"),
               avg_val=("valuation_b", "mean"),
               total_val=("valuation_b", "sum"))
          .dropna().reset_index().sort_values("year_joined").round(2))
pivot5["year_joined"] = pivot5["year_joined"].astype(int)
pivot5["cumulative_total"] = pivot5["new_unicorns"].cumsum()
pivot5["yoy_growth_pct"]   = pivot5["new_unicorns"].pct_change().mul(100).round(1)
pivot5.columns = [
    "Year", "New Unicorns", "Avg Valuation ($B)",
    "Total Valuation ($B)", "Cumulative Total", "YoY Growth %",
]

# ── Write sheets ──────────────────────────────────────────────────────────────
dfs = {
    "Industry × Continent":  pivot1,
    "Valuation Tiers":        pivot2,
    "Decade × Industry":      pivot3,
    "Country League Table":   pivot4,
    "Yearly Growth":          pivot5,
}

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    for sheet_name, data in dfs.items():
        data.to_excel(writer, sheet_name=sheet_name, startrow=2, index=True
                      if sheet_name in ("Industry × Continent", "Decade × Industry")
                      else False)
        print(f"  Written: {sheet_name}")

# ── Style workbook ────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_PATH)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.freeze_panes  = "B4" if sheet_name in ("Industry × Continent",
                                               "Decade × Industry") else "A3"
    ws.sheet_view.showGridLines = False

    add_title(ws, sheet_name, row=1)

    if sheet_name == "Industry × Continent":
        n_cols = ws.max_column
        style_header_row(ws, row=3, n_cols=n_cols, bg=DARK_GREEN)
        style_index_col(ws, start_row=4, end_row=ws.max_row - 1, col=1)
        # TOTAL row
        total_row = ws.max_row
        for col in range(1, n_cols + 1):
            ws.cell(row=total_row, column=col).fill = make_fill(MID_GREEN)
            ws.cell(row=total_row, column=col).font = make_font(bold=True,
                                                                color=WHITE)
        # Colour scale on data cells
        data_range = (f"B4:{get_column_letter(n_cols - 1)}"
                      f"{total_row - 1}")
        ws.conditional_formatting.add(data_range, ColorScaleRule(
            start_type="min",  start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="D8F3DC",
            end_type="max",    end_color="1A3C34",
        ))
        zebra_rows(ws, 4, total_row - 1, 2, n_cols)

    elif sheet_name == "Valuation Tiers":
        n_cols = pivot2.shape[1]
        style_header_row(ws, row=3, n_cols=n_cols, bg=MID_GREEN)
        zebra_rows(ws, 4, ws.max_row, 1, n_cols)
        # Data bar on # Companies column (col B)
        ws.conditional_formatting.add(
            f"B4:B{ws.max_row}",
            DataBarRule(start_type="min", end_type="max",
                        color=ACCENT_GREEN),
        )

    elif sheet_name == "Decade × Industry":
        n_cols = ws.max_column
        style_header_row(ws, row=3, n_cols=n_cols, bg=DARK_GREEN)
        style_index_col(ws, start_row=4, end_row=ws.max_row, col=1)
        data_range = f"B4:{get_column_letter(n_cols)}{ws.max_row}"
        ws.conditional_formatting.add(data_range, ColorScaleRule(
            start_type="min",  start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="D8F3DC",
            end_type="max",    end_color="2D6A4F",
        ))

    elif sheet_name == "Country League Table":
        n_cols = pivot4.shape[1]
        style_header_row(ws, row=3, n_cols=n_cols, bg=MID_GREEN)
        zebra_rows(ws, 4, ws.max_row, 1, n_cols)
        # Highlight top 3
        for row in range(4, 7):
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = make_fill(ACCENT_GREEN)
                ws.cell(row=row, column=col).font = make_font(bold=True,
                                                              color=WHITE)
        # Data bar on # Unicorns (col D)
        ws.conditional_formatting.add(
            f"D4:D{ws.max_row}",
            DataBarRule(start_type="min", end_type="max",
                        color=MID_GREEN),
        )

    elif sheet_name == "Yearly Growth":
        n_cols = pivot5.shape[1]
        style_header_row(ws, row=3, n_cols=n_cols, bg=MID_GREEN)
        zebra_rows(ws, 4, ws.max_row, 1, n_cols)
        # Colour scale on cumulative total (col E)
        ws.conditional_formatting.add(
            f"E4:E{ws.max_row}",
            ColorScaleRule(
                start_type="min",  start_color="D8F3DC",
                end_type="max",    end_color="1A3C34",
            ),
        )

    # Row heights and autofit for all sheets
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 6   # spacer
    ws.row_dimensions[3].height = 28  # header
    for row_idx in range(4, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18
    autofit(ws)

wb.save(OUTPUT_PATH)
print(f"\nSaved → {OUTPUT_PATH}")
print(f"Sheets: {', '.join(wb.sheetnames)}")